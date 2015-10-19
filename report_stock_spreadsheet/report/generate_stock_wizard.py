#from openerp.osv import fields, osv, orm


from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.worksheet import ColumnDimension
from openpyxl.styles import Border, Side, Font, Style
import base64
from openerp.tools.translate import _
from openerp import models,fields,api,exceptions,tools


class generate_stock_wizard(models.Model):
    _name = "report_stock.generate_stock"
    data=fields.Binary('File', readonly=True)
    name=fields.Char('File Name', readonly=True)
    state=fields.Selection([('choose', 'choose'),
                                   ('get', 'get')],default='choose')
    
    def generate_file(self, cr, uid, ids, context=None):

        if context is None:
            context = {}
        
        product_obj = self.pool.get('product.product')
        quant_obj = self.pool.get('stock.quant')
        location_obj = self.pool.get('stock.location')
                
        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()

        wb = Workbook()

        ws = wb.active

        ws.title = "Stock_Total"
        ws['A1'].value = "Report Total Stock"

        ws['A2'].value = "id"
        ws['B2'].value = "default_code"
        ws['C2'].value = "ean14"
        ws['D2'].value = "category_name"
        ws['E2'].value = "name"
        ws['F2'].value = "cost"
        ws['G2'].value = "price"
        ws['H2'].value = "stock_total"
        ws['I2'].value = "costo_total"
        ws.merge_cells('A1:I1')
#        ws.freeze_panes = 'A2'  no funciona 
        
        
        border_bottom = Border(bottom=Side(style='thin'))
        border_right = Border(right=Side(style='thin'))
        border_corner = Border(right=Side(style='thin'),bottom=Side(style='thin'))
        font_bold = Font(bold=True)

        sql = """
            Select product_id,sum(qty) qty
                from stock_quant sq JOIN
                stock_location sl ON sq.location_id = sl.id
                where sl.usage = 'internal'
                group by product_id
         """   
#                 AND pp.default_code = 'RVR022C-NEK'

        cr.execute(sql,)
        row = 3
        for product_line in cr.dictfetchall():

            product_id = product_obj.browse(cr, uid,product_line['product_id'] )

            ws.cell(row=row, column= 1).value = product_id.id
            ws.cell(row=row, column= 2).value = product_id.default_code
            ws.cell(row=row, column= 3).value = product_id.ean13
            ws.cell(row=row, column= 4).value = product_id.categ_id.name
            ws.cell(row=row, column= 5).value = product_id.name
            ws.cell(row=row, column= 6).value = product_id.standard_price
            ws.cell(row=row, column= 7).value = product_id.list_price
            ws.cell(row=row, column= 8).value = product_line['qty']
            ws.cell(row=row, column= 9).value = (product_line['qty']*product_id.standard_price)
            
            ws.cell(row=row, column= 1).border = border_right
            ws.cell(row=row, column= 2).border = border_right
            ws.cell(row=row, column= 5).border = border_right
            ws.cell(row=row, column= 6).border = border_right
            ws.cell(row=row, column= 7).border = border_right
            ws.cell(row=row, column= 8).border = border_right
            ws.cell(row=row, column= 9).border = border_right
            row += 1


        for r in ws.iter_rows('A1:G1'):
            for c in r:
                c.font = font_bold
        for r in ws.iter_rows('A2:G2'):
            for c in r:
                c.border = border_bottom
                c.font = font_bold


        ws['A1'].border = border_right
        ws['A2'].border = border_corner
        ws['B1'].border = border_right
        ws['B2'].border = border_corner
        ws['C1'].border = border_right
        ws['C2'].border = border_corner
        ws['D1'].border = border_right
        ws['D2'].border = border_corner
        ws['E1'].border = border_right
        ws['E2'].border = border_corner
        ws['F1'].border = border_right
        ws['F2'].border = border_corner
        ws['G1'].border = border_right
        ws['G2'].border = border_corner
        ws['H1'].border = border_right
        ws['H2'].border = border_corner
        ws['I1'].border = border_right
        ws['I2'].border = border_corner

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['E'].width = 70

        
        # 
        # Location = Iternal
        #

        location_ids = location_obj.search(cr, uid, [('usage', '=', 'internal')], context=context)
        for location in location_ids:

            location_id = location_obj.browse(cr, uid,location)
            print location_id.display_name
            ws = wb.create_sheet()
            ws['A1'].value = location_id.display_name
            ws.title = location_id.display_name.replace('/', '_')
            ws['A2'].value = "id"
            ws['B2'].value = "default_code"
            ws['C2'].value = "ean14"
            ws['D2'].value = "category_name"
            ws['E2'].value = "name"
            ws['F2'].value = "cost"
            ws['G2'].value = "price"
            ws['H2'].value = "stock_total"
            ws['I2'].value = "costo_total"
            ws.merge_cells('A1:I1')
            
            sql_quant = """
                Select product_id,location_id,sum(qty) qty
                from stock_quant
            """   
            cr.execute(sql_quant + "Where location_id = %s group by product_id,location_id", (location,))

            row = 3
            for product_quant in cr.dictfetchall():
                product_id = product_obj.browse(cr, uid,product_quant['product_id'] )

                # se implemento un query a quant para que sea mas eficiente la extraccion de datos de la DB
                #product_quant_ids = quant_obj.search(cr, uid, [('location_id', '=', location),('product_id', '=', product_id.id)], context=context)
                #for quant in product_quant_ids:
                #    quant_id = quant_obj.browse(cr, uid,quant)
                #    product_stock+=quant_id.qty         

                ws.cell(row=row, column= 1).value = product_id.id
                ws.cell(row=row, column= 2).value = product_id.default_code
                ws.cell(row=row, column= 3).value = product_id.ean13
                ws.cell(row=row, column= 4).value = product_id.categ_id.name
                ws.cell(row=row, column= 5).value = product_id.name
                ws.cell(row=row, column= 6).value = product_id.standard_price
                ws.cell(row=row, column= 7).value = product_id.list_price
                ws.cell(row=row, column= 8).value = product_quant['qty']
                ws.cell(row=row, column= 9).value = (product_quant['qty']*product_id.standard_price)
            
                ws.cell(row=row, column= 1).border = border_right
                ws.cell(row=row, column= 2).border = border_right
                ws.cell(row=row, column= 5).border = border_right
                ws.cell(row=row, column= 6).border = border_right
                ws.cell(row=row, column= 7).border = border_right
                ws.cell(row=row, column= 8).border = border_right
                ws.cell(row=row, column= 9).border = border_right
                ws.cell(row=row, column= 10).border = border_right

                row += 1

            for r in ws.iter_rows('A1:J1'):
                for c in r:
                    c.font = font_bold
            for r in ws.iter_rows('A2:J2'):
                for c in r:
                    c.border = border_bottom
                    c.font = font_bold


            ws['A1'].border = border_right
            ws['A2'].border = border_corner
            ws['B1'].border = border_right
            ws['B2'].border = border_corner
            ws['C1'].border = border_right
            ws['C2'].border = border_corner
            ws['D1'].border = border_right
            ws['D2'].border = border_corner
            ws['E1'].border = border_right
            ws['E2'].border = border_corner
            ws['F1'].border = border_right
            ws['F2'].border = border_corner
            ws['G1'].border = border_right
            ws['G2'].border = border_corner
            ws['H1'].border = border_right
            ws['H2'].border = border_corner
            ws['I1'].border = border_right
            ws['I2'].border = border_corner
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['E'].width = 70

        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)


        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "SBG_Product_stock.xlsx",
            'data': out
        }, context=context)
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'report_stock.generate_stock',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

